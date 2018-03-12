import json
import openpyxl

jsondata = open("路外停車資訊.json", 'r', encoding='UTF-8')

data = jsondata.read()
data = json.loads(data)

#宣告一個試算表
workbook = openpyxl.Workbook()
workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))
workbook.create_sheet('桃園市停車場資訊')

#操作一個工作表
sheet = workbook.get_sheet_by_name('桃園市停車場資訊')

#print (len(data['parkingLots']))

#寫入值
sheet['A1'] = 'areaId'
sheet['B1'] = 'areaName'
sheet['C1'] = 'parkName'
sheet['D1'] = 'totalSpace'
sheet['E1'] = 'surplusSpace'
sheet['F1'] = 'payGuide'
sheet['G1'] = 'introduction'
sheet['H1'] = 'address'
sheet['I1'] = 'wgsX'
sheet['J1'] = 'wgsY'
sheet['K1'] = 'parkId'
x = 0
while x < len(data['parkingLots']):
    sheet.cell(row = x+2, column = 1).value = data['parkingLots'][x]['areaId']
    sheet.cell(row = x+2, column = 2).value = data['parkingLots'][x]['areaName']
    sheet.cell(row = x+2, column = 3).value = data['parkingLots'][x]['parkName']
    sheet.cell(row = x+2, column = 4).value = data['parkingLots'][x]['totalSpace']
    sheet.cell(row = x+2, column = 5).value = data['parkingLots'][x]['surplusSpace']
    sheet.cell(row = x+2, column = 6).value = data['parkingLots'][x]['payGuide']
    sheet.cell(row = x+2, column = 7).value = data['parkingLots'][x]['introduction']
    sheet.cell(row = x+2, column = 8).value = data['parkingLots'][x]['address']
    sheet.cell(row = x+2, column = 9).value = data['parkingLots'][x]['wgsX']
    sheet.cell(row = x+2, column = 10).value = data['parkingLots'][x]['wgsY']
    sheet.cell(row = x+2, column = 11).value = data['parkingLots'][x]['parkId']
    x+=1

#記得存檔歐
workbook.save('test1.xlsx')


    

